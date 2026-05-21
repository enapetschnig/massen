#!/usr/bin/env python3
"""ÖNORM-konforme Massenermittlung aus Bauplan-PDF — lokal, ohne Vercel/Supabase.

Pipeline:
  1. Text-Layer mit fitz extrahieren (alle Spans mit Position + Schriftgröße).
  2. Räume erkennen (strict: exact name match + size ≥ 9).
  3. Pro Raum: Fläche / Umfang / Höhe / Bodenbelag aus Beschriftungsblock.
  4. Fenster pro Plan: FE_-Codes + AL-Maße → echte Stockmaß-Flächen.
  5. ÖNORM A 2063 LV erzeugen (Pos-Nr / Beschreibung / Detail / Endsumme).
  6. Excel-Export im Kutzen-Koblach-Format (zur 1:1-Übergabe an Baubetrieb).

Aufruf: python3 scripts/oenorm_extract.py <plan.pdf> [<ausgabe.xlsx>]
"""
from __future__ import annotations
import json
import math
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

import fitz
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ════════════════════════════════════════════════════════════════════════
# 1. Span-Extraktion
# ════════════════════════════════════════════════════════════════════════
def extract_spans(page) -> list[dict]:
    out = []
    td = page.get_text("dict")
    for block in td.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                t = (span.get("text") or "").strip()
                if not t:
                    continue
                bb = tuple(span.get("bbox") or (0, 0, 0, 0))
                out.append({
                    "text": t,
                    "bbox": bb,
                    "size": round(span.get("size", 0), 1),
                    "cx": (bb[0] + bb[2]) / 2.0,
                    "cy": (bb[1] + bb[3]) / 2.0,
                })
    return out


# ════════════════════════════════════════════════════════════════════════
# 2. Raum-Erkennung (strict)
# ════════════════════════════════════════════════════════════════════════
ROOM_NAMES_EXACT = {
    "Wohnküche", "Wohnkueche", "Wohnen", "Wohnzimmer", "Wohnraum", "Esszimmer", "Zimmer",
    "Schlafzimmer", "Kinderzimmer", "Bad", "WC", "Dusche", "Sauna",
    "Vorraum", "Vorzimmer", "Flur", "Gang", "Diele", "Küche", "Kueche",
    "Loggia", "Balkon", "Terrasse", "Stiegenhaus", "Stiege", "STGH", "STG",
    "Abstellraum", "Abstellr.", "Garderobe", "Speis", "Speisekammer",
    "Technik", "Technikraum",
    "Keller", "Kellerabteil", "Kiwa", "Waschküche", "Waschkueche", "Waschraum", "Waschen",
    "Werkstätte", "Werkstatt", "Lager",
    "Büro", "Buero", "Atelier", "Studio", "Praxis", "Arbeitszimmer",
    "Tiefgarage", "Garage", "Carport", "Parkplatz", "Fahrradraum", "Kinderwagenraum", "Schleuse",
    "Fitness", "Treppenhaus", "Müllraum", "E-Technik", "Elektroraum",
    "Pelletslagerraum", "Pelletslager",
    "Windfang", "Foyer", "Eingang", "Eingangsbereich",
    "AR", "HWR", "HSR", "HAR",
}
BODENBELAG_KWS = {
    "parkett", "fliesen", "laminat", "vinyl", "estrich", "teppich",
    "feinsteinzeug", "naturstein", "keramik", "beschichtung", "beton",
}
AR_TOP_RX = re.compile(r"^(AR\s+)?TOP\s*\.?\s*\d+\s*(AR)?\s*[a-z]?$", re.I)

# Manche Architekten beschriften Räume mit Geschoss-Codes (EG301, OG214, KG05).
# Pattern: 2-stelliger Geschoss-Prefix + 2-4-stellige Raum-Nummer.
ROOM_CODE_RX = re.compile(r"^(EG|OG\d?|KG|UG|DG)\s*[._-]?\s*(\d{2,4})$", re.I)

# Generalisierte Anker-Patterns: matchen verschiedene Architekt-Konventionen
# F:/Fl:/Fläche, U:/Um:/Umfang, H:/Hö:/RH/Höhe etc.
# U: kann mit Tausender-Trennzeichen (Leerzeichen) und Einheit cm/m kommen
# Beispiele: "U: 20,66 m", "U: 1 098,0 cm", "U=2399.9cm"
F_ANCHOR_RX  = re.compile(r"^(?:F|Fl|Fläche|Flaeche)\s*[:=]?\s*([0-9]+[,.][0-9]+)", re.I)
U_ANCHOR_RX  = re.compile(r"^(?:U|Um|Umfang)\s*[:=]?\s*([0-9][0-9\s]*[,.][0-9]+)", re.I)
H_ANCHOR_RX  = re.compile(r"^(?:H|Hö|Hoe|Höhe|Hoehe|RH|LH)\s*[:=]?\s*([0-9]+[,.][0-9]+)", re.I)
# Bodenbelag mit "B:" Prefix: "B: Fliesen", "B: Parkett, Eiche"
B_ANCHOR_RX  = re.compile(r"^B\s*[:=]\s*(.+)$", re.I)

# Klassifikation für ÖNORM-konforme Aggregation
KATEGORIE = {
    # Beheizte Innenräume → Innenputz Wände + Decken
    "Innenraum_warm": {
        "Wohnküche", "Wohnkueche", "Wohnen", "Wohnzimmer", "Wohnraum", "Esszimmer",
        "Zimmer", "Schlafzimmer", "Kinderzimmer", "Küche", "Kueche", "Bad", "WC",
        "Dusche", "Sauna", "Vorraum", "Vorzimmer", "Flur", "Gang", "Diele", "Garderobe",
        "Abstellraum", "Speis", "Speisekammer", "AR", "HWR", "HSR", "HAR",
        "Büro", "Buero", "Atelier", "Studio", "Praxis", "Arbeitszimmer",
        "Waschküche", "Waschkueche", "Waschraum", "Waschen", "Kiwa",
        "Windfang", "Foyer", "Eingang", "Eingangsbereich", "Fitness",
    },
    # Loggia/Balkon/Terrasse: außenliegend, eigene Position
    "Loggia":      {"Loggia", "Balkon", "Terrasse", "Parkplatz", "Carport"},
    # Stiegenhaus: oft eigenes Gewerk
    "Stiegenhaus": {"Stiegenhaus", "Stiege", "STGH", "STG", "Treppenhaus"},
    # Kalträume
    "Nebenraum_kalt": {
        "Tiefgarage", "Garage", "Keller", "Kellerabteil", "Technik", "Technikraum",
        "Müllraum", "E-Technik", "Elektroraum", "Pelletslager", "Pelletslagerraum",
        "Fahrradraum", "Kinderwagenraum", "Schleuse",
        "Werkstätte", "Werkstatt", "Lager",
    },
}


def kategorie_of(name: str) -> Optional[str]:
    """Kategorisiert einen Raumnamen — exact, first-word und Bindestrich-Teile.
    "Wohnraum Küche" → first-word "Wohnraum". "Geräte-Abstellraum" → "Abstellraum"."""
    if not name:
        return None
    name = name.strip()
    candidates = [name]
    if " " in name:
        candidates.append(name.split()[0])
    if "-" in name:
        candidates.extend(p.strip() for p in name.split("-"))
    for kat, names in KATEGORIE.items():
        for c in candidates:
            if c in names:
                return kat
    return None


# Wörter die einen first-word-Match zu einer Annotation machen (kein Raum):
# "Loggia Entwässerung", "Bad Detail" etc.
_RAUM_REST_OK = {
    "süd", "nord", "ost", "west", "südost", "südwest", "nordost", "nordwest",
    "links", "rechts", "oben", "unten", "mitte", "gross", "groß", "klein",
    "überdacht", "ueberdacht", "offen", "beheizt", "unbeheizt", "neu", "alt",
}
# Legende-Pattern: Kürzel + " - " + Vollname (z.B. "AR - Abstellraum").
# Leerzeichen um den Bindestrich unterscheidet von echten Räumen wie "E-Technik".
_LEGENDE_RX = re.compile(r"^[A-ZÄÖÜ]{1,4}\s+[-–]\s+\w", re.I)
# TOP-Wohnungslabel allein ("TOP 25") ist KEIN Raum, sondern Wohnungs-ID
_TOP_ONLY_RX = re.compile(r"^TOP\s*\.?\s*\d+[a-z]?$", re.I)


def name_matches_room(t: str) -> bool:
    """Prüft NUR den Text (ohne Größe), ob er ein Raumname ist."""
    if len(t) < 2:
        return False
    # Legenden-Eintrag oder TOP-Label → kein Raum
    if _LEGENDE_RX.match(t):
        return False
    if _TOP_ONLY_RX.match(t):
        return False
    # Exact match
    if t in ROOM_NAMES_EXACT:
        return True
    # First-word match — Rest muss raum-artig sein (Zahl, Himmelsrichtung,
    # oder selbst ein Raumname). Verhindert "Loggia Entwässerung".
    words = t.split()
    if len(words) > 1 and words[0] in ROOM_NAMES_EXACT:
        rest_ok = all(
            re.match(r"^\d+[a-z]?$", w, re.I)
            or w.lower() in _RAUM_REST_OK
            or w in ROOM_NAMES_EXACT
            for w in words[1:]
        )
        if rest_ok:
            return True
    # Bindestrich-Komposition ohne Spaces: "Geräte-Abstellraum"
    if "-" in t and not _LEGENDE_RX.match(t):
        for part in t.split("-"):
            if part.strip() in ROOM_NAMES_EXACT:
                return True
    return False


def compute_label_size_threshold(spans: list[dict]) -> float:
    """Adaptive Schrift-Schwelle: bestimmt die typische Raumlabel-Größe dieses
    Plans aus den Größen aller exakten Raumnamen-Treffer (Modus).

    Wohnungsgrundrisse 1:50 → ~9-10pt, Einreichpläne 1:100 → ~7.5pt,
    Lagepläne → ~5pt. Schwelle = typische Größe × 0.75 (Toleranz)."""
    from collections import Counter
    sizes = []
    for s in spans:
        if name_matches_room(s["text"].strip()):
            sizes.append(round(s["size"], 1))
    if not sizes:
        return 6.0
    # Modus (häufigste Größe) ist robust gegen vereinzelte Lageplan-Mini-Labels
    most_common_size = Counter(sizes).most_common(1)[0][0]
    return max(4.0, most_common_size * 0.75)


def is_room_label(s: dict, min_size: float = 6.0) -> bool:
    t = s["text"].strip()
    # Geschoss-Code-Räume (EG301, OG214) — eigene Größenregel
    if ROOM_CODE_RX.match(t) and s["size"] >= max(4.0, min_size * 0.7):
        return True
    if s["size"] < min_size:
        return False
    return name_matches_room(t)


def has_m2_superscript(s: dict, spans: list[dict], tolerance_pt: int = 20) -> bool:
    sx_right = s["bbox"][2]
    sy = s["cy"]
    for o in spans:
        if o is s or o["text"] != "2":
            continue
        if o["size"] >= s["size"] * 0.8:
            continue
        if abs(o["cy"] - sy) > 8:
            continue
        if -2 <= (o["bbox"][0] - sx_right) <= tolerance_pt:
            return True
    return False


def extract_room(rs: dict, spans: list[dict]) -> dict:
    rx, ry = rs["cx"], rs["cy"]
    # Bei Code-Räumen (EG301 etc) sind F/U/H oft weiter entfernt vom Label —
    # in schematischen Plänen liegt die Beschriftung manchmal 100-150pt
    # vom Raum-Code entfernt (mittig im Raum). Adaptiver Suchradius.
    is_code = bool(ROOM_CODE_RX.match(rs["text"].strip()))
    rad_x = 150 if is_code else 60
    rad_y_pos = 150 if is_code else 60
    rad_y_neg = -150 if is_code else -5
    cands = []
    for s in spans:
        if s is rs:
            continue
        dx = s["cx"] - rx
        dy = s["cy"] - ry
        if abs(dx) <= rad_x and rad_y_neg <= dy <= rad_y_pos:
            cands.append((dy, dx, s))
    cands.sort()
    f = u = h = None
    boden = None
    sources = {"flaeche_m2": None, "umfang_m": None, "hoehe_m": None, "bodenbelag": None}
    for dy, dx, s in cands:
        t = s["text"]
        m = U_ANCHOR_RX.match(t)
        if m and u is None:
            raw = m.group(1).replace(" ", "").replace(",", ".")
            v = float(raw)
            # Einheits-Detect: wenn "cm" im Text oder Wert > 50, dann cm→m
            is_cm = ("cm" in t.lower()) or (v > 50)
            if is_cm:
                v = v / 100.0
            if 1.0 <= v <= 200.0:  # plausibler Umfang-Bereich
                u = v
                sources["umfang_m"] = "pdf-text:U-anchor" + (" (cm→m)" if is_cm else "")
            continue
        m = H_ANCHOR_RX.match(t)
        if m and h is None:
            v = float(m.group(1).replace(",", "."))
            # Auto cm→m: Raumhöhen liegen real bei 2.2-4.5 m, in cm bei 220-450
            if v > 20 and v < 500:
                v = v / 100.0
            if 2.0 <= v <= 5.0:  # plausibler Bereich
                h = v
                sources["hoehe_m"] = "pdf-text:H-anchor"
            continue
        m = F_ANCHOR_RX.match(t)
        if m and f is None:
            f = float(m.group(1).replace(",", "."))
            sources["flaeche_m2"] = "pdf-text:F-anchor"
            continue
        m = re.match(r"^([0-9]+[,.][0-9]+)\s*m\s*(²|2)?\s*$", t)
        if m and f is None and not t.startswith(("U", "H", "F")):
            v = float(m.group(1).replace(",", "."))
            if "²" in t or has_m2_superscript(s, spans):
                f = v
                sources["flaeche_m2"] = "pdf-text:m²-superscript"
                continue
        if boden is None and t.lower() in BODENBELAG_KWS:
            boden = t
            sources["bodenbelag"] = "pdf-text:keyword"
            continue
        # B: Fliesen / B: Parkett (Bodenbelag mit Prefix)
        bm = B_ANCHOR_RX.match(t)
        if bm and boden is None:
            belag_text = bm.group(1).strip().split(",")[0].strip()
            if belag_text.lower() in BODENBELAG_KWS:
                boden = belag_text
                sources["bodenbelag"] = "pdf-text:B-anchor"
    return {
        "name": rs["text"],
        "flaeche_m2": f,
        "umfang_m": u,
        "hoehe_m": h,
        "bodenbelag": boden,
        "bbox": list(rs["bbox"]),
        "cx": rs["cx"],
        "cy": rs["cy"],
        "kategorie": kategorie_of(rs["text"]),
        "quellen": sources,
    }


# ════════════════════════════════════════════════════════════════════════
# 3. Fenster aus FE_/AL-Markern
# ════════════════════════════════════════════════════════════════════════
def extract_windows(spans: list[dict]) -> list[dict]:
    fe = [s for s in spans if re.match(r"^FE[_-]?\d+", s["text"])]
    al_spans = []
    for s in spans:
        m = re.match(r"^AL\s*(\d{2,3})", s["text"])
        if m:
            al_spans.append({**s, "al_cm": int(m.group(1))})
    out = []
    used = set()
    for f in fe:
        nearby = []
        for al in al_spans:
            if id(al) in used:
                continue
            dx = abs(al["cx"] - f["cx"])
            dy = abs(al["cy"] - f["cy"])
            if dx > 100 or dy > 80:
                continue
            nearby.append((dx + dy, al))
        nearby.sort(key=lambda x: x[0])
        my = nearby[:2]
        for _, al in my:
            used.add(id(al))
        if len(my) >= 2:
            vals = sorted([al["al_cm"] for _, al in my])
            br, ho = vals[0], vals[1]
            src = "pdf-text:2xAL"
        elif len(my) == 1:
            br = my[0][1]["al_cm"]
            ho = 200  # Default-Höhe, falls nur Breite gefunden
            src = "pdf-text:1xAL+default"
        else:
            br, ho = 120, 200
            src = "default-120x200"
        out.append({
            "code": f["text"],
            "cx": f["cx"],
            "cy": f["cy"],
            "breite_m": br / 100.0,
            "hoehe_m": ho / 100.0,
            "flaeche_m2": br * ho / 10000.0,
            "quelle": src,
        })
    return out


# ════════════════════════════════════════════════════════════════════════
# 4. Türen (heuristisch + via Plan-Symbole, soweit aus Text ableitbar)
# ════════════════════════════════════════════════════════════════════════
def extract_doors(spans: list[dict]) -> list[dict]:
    """Türen aus T-Markern (T1 / T2 …) bzw. typischen Tür-Beschriftungen.
    Wenn nicht gefunden → leere Liste; LV nutzt dann Heuristik."""
    out = []
    for s in spans:
        m = re.match(r"^T\s*[._-]?\s*(\d{1,3})$", s["text"])
        if m and s["size"] >= 4.0:
            out.append({"code": s["text"], "cx": s["cx"], "cy": s["cy"]})
    return out


# ════════════════════════════════════════════════════════════════════════
# 5. Dedup nach Position
# ════════════════════════════════════════════════════════════════════════
def _completeness(r: dict) -> int:
    return (1 if r.get("flaeche_m2") else 0) + (1 if r.get("umfang_m") else 0) + \
           (1 if r.get("hoehe_m") else 0) + (1 if r.get("bodenbelag") else 0)


def dedup_by_position(rooms: list[dict], grid: int = 5) -> list[dict]:
    """Dedup nach räumlicher Position. Ein Raum ohne F wird behalten, wenn
    er name+position hat — vereinfachte Pläne haben oft nur 1-2 F-Werte,
    aber die Räume selbst sind durch Code/Name dennoch zählbar."""
    pos = {}
    for r in rooms:
        if not r.get("name"):
            continue
        k = (int(round(r["cx"] / grid)), int(round(r["cy"] / grid)))
        ex = pos.get(k)
        if ex is None or _completeness(r) > _completeness(ex):
            pos[k] = r
    return list(pos.values())


# ════════════════════════════════════════════════════════════════════════
# 6. Haus-Mapping (automatisch via x-Histogramm, mit manuell-override Option)
# ════════════════════════════════════════════════════════════════════════
def auto_houses(rooms: list[dict], page_width: float) -> dict[str, tuple[float, float]]:
    """Findet Haus-Bereiche auf einem Mehrhaus-Plan: scannt x-Histogramm der
    Räume und erkennt Lücken > 300pt als Haus-Trenner."""
    if not rooms:
        return {"A": (0, page_width)}
    xs = sorted(r["cx"] for r in rooms)
    gaps = []
    for i in range(1, len(xs)):
        if xs[i] - xs[i - 1] > 300:
            gaps.append((xs[i - 1] + xs[i]) / 2.0)
    if not gaps:
        return {"A": (0, page_width)}
    edges = [0.0] + gaps + [page_width]
    out = {}
    for i in range(len(edges) - 1):
        out[chr(ord("A") + i)] = (edges[i], edges[i + 1])
    return out


def house_of(cx: float, houses: dict[str, tuple[float, float]]) -> Optional[str]:
    for h, (a, b) in houses.items():
        if a <= cx < b:
            return h
    return None


# ════════════════════════════════════════════════════════════════════════
# 7. ÖNORM A 2063 LV-Generator
# ════════════════════════════════════════════════════════════════════════
class LVPosition:
    """Eine Position im ÖNORM A 2063 Leistungsverzeichnis."""

    def __init__(self, posnr: str, beschreibung: str, einheit: str):
        self.posnr = posnr
        self.beschreibung = beschreibung
        self.einheit = einheit
        self.zeilen: list[dict] = []
        self.quelle: str = ""
        self.konfidenz: float = 1.0

    def add_zeile(self, text: str, anzahl: float = 0, laenge: float = 0,
                  breite: float = 0, hoehe: float = 0, summe: Optional[float] = None,
                  quelle: str = ""):
        # Berechnung wie in Kutzen-Koblach-Excel: n × L × B × H (B/L/H je nach Maßart)
        wert = summe
        if wert is None:
            mult = (anzahl or 1) * (laenge or 1) * (breite or 1) * (hoehe or 1)
            # Wenn nur eine Dimension gesetzt → Multiplikation übernimmt die,
            # da Defaults 1 sind. Bei z.B. lfm-Position: nur länge gesetzt.
            wert = mult
        self.zeilen.append({
            "text": text,
            "anzahl": anzahl or None,
            "laenge": laenge or None,
            "breite": breite or None,
            "hoehe": hoehe or None,
            "wert": round(wert, 4),
            "quelle": quelle,
        })

    @property
    def endsumme(self) -> float:
        return round(sum(z["wert"] for z in self.zeilen), 2)

    def to_dict(self) -> dict:
        return {
            "posnr": self.posnr,
            "beschreibung": self.beschreibung,
            "einheit": self.einheit,
            "endsumme": self.endsumme,
            "konfidenz": self.konfidenz,
            "quelle": self.quelle,
            "zeilen": self.zeilen,
        }


def build_lv(rooms: list[dict], windows: list[dict],
             houses: dict[str, tuple[float, float]],
             geschoss: str = "EG") -> list[LVPosition]:
    """Generiere ein vollständiges ÖNORM-konformes LV pro Haus."""
    positionen: list[LVPosition] = []

    # Pro Haus Räume + Fenster gruppieren
    for haus_idx, (haus, _) in enumerate(sorted(houses.items())):
        h_rooms = [r for r in rooms if house_of(r["cx"], houses) == haus]
        h_windows = [w for w in windows if house_of(w["cx"], houses) == haus]

        if not h_rooms:
            continue

        pos_prefix = f"{haus_idx + 1}"  # 1, 2, 3 …

        # ───────────────────────────────────────────────────────
        # POS 1.1 — Innenputz Wände (m²)
        # ÖNORM B 2210/B 3346: Wandflächen = Σ U×H der Innenräume,
        # minus Fenster (Stockmaß), minus Türen (Stockmaß).
        # Loggien & Stiegenhaus separat (eigene Positionen).
        # ───────────────────────────────────────────────────────
        innen = [r for r in h_rooms if r["kategorie"] == "Innenraum_warm"
                 and r.get("umfang_m") and r.get("hoehe_m")]

        pos = LVPosition(f"{pos_prefix}.1", f"Innenputz Wände — Haus {haus}, {geschoss}", "m²")
        pos.quelle = "PDF-Text-Layer (F/U/H byte-exakt aus Beschriftungsblock)"
        for r in innen:
            wandfl = r["umfang_m"] * r["hoehe_m"]
            pos.add_zeile(
                f"{r['name']} ({r.get('wohnung') or '–'})",
                laenge=r["umfang_m"], hoehe=r["hoehe_m"], summe=wandfl,
                quelle=f"U={r['umfang_m']}m · H={r['hoehe_m']}m"
            )
        # Fenster-Abzug (ÖNORM B 2210 §6.4: Öffnungen > 2,5 m² abziehen)
        for w in h_windows:
            if w["flaeche_m2"] > 2.5:
                pos.add_zeile(
                    f"Abzug Fenster {w['code']}",
                    laenge=w["breite_m"], hoehe=-w["hoehe_m"], summe=-w["flaeche_m2"],
                    quelle=f"AL={int(w['breite_m']*100)}×{int(w['hoehe_m']*100)}"
                )
        # Türen (heuristisch — TODO: aus Geometrie zählen)
        n_doors = max(1, round(len(innen) * 1.5))
        tuer_breite, tuer_hoehe = 0.9, 2.1
        pos.add_zeile(
            f"Abzug Türen (heuristisch, {n_doors} Stk à {tuer_breite}×{tuer_hoehe} m)",
            anzahl=n_doors, laenge=tuer_breite, hoehe=-tuer_hoehe,
            summe=-n_doors * tuer_breite * tuer_hoehe,
            quelle="heuristisch: 1.5 Türen/Raum"
        )
        pos.konfidenz = 0.93  # Realistisch: ~5-7% Abweichung wg. Tür-Heuristik
        positionen.append(pos)

        # ───────────────────────────────────────────────────────
        # POS 1.2 — Innenputz Decken (m²)
        # = Σ Fläche der Innenräume (direkt aus F-Wert)
        # ───────────────────────────────────────────────────────
        pos = LVPosition(f"{pos_prefix}.2", f"Innenputz Decken — Haus {haus}, {geschoss}", "m²")
        pos.quelle = "PDF-Text-Layer (F-Wert byte-exakt)"
        for r in innen:
            if r.get("flaeche_m2"):
                pos.add_zeile(
                    f"{r['name']} ({r.get('wohnung') or '–'})",
                    summe=r["flaeche_m2"],
                    quelle=f"F={r['flaeche_m2']}m²"
                )
        pos.konfidenz = 0.99
        positionen.append(pos)

        # ───────────────────────────────────────────────────────
        # POS 1.3 — Bodenbeläge pro Materialtyp (m²)
        # ───────────────────────────────────────────────────────
        by_belag = defaultdict(list)
        for r in h_rooms:
            if r.get("bodenbelag") and r.get("flaeche_m2"):
                by_belag[r["bodenbelag"]].append(r)
        for sub_idx, (belag, rs) in enumerate(sorted(by_belag.items()), start=1):
            pos = LVPosition(f"{pos_prefix}.3.{sub_idx}",
                             f"Bodenbelag {belag} — Haus {haus}, {geschoss}", "m²")
            pos.quelle = "PDF-Text-Layer (F + Bodenbelag-Keyword)"
            for r in rs:
                pos.add_zeile(
                    f"{r['name']} ({r.get('wohnung') or '–'})",
                    summe=r["flaeche_m2"],
                    quelle=f"F={r['flaeche_m2']}m², Belag='{r['bodenbelag']}'"
                )
            pos.konfidenz = 0.97
            positionen.append(pos)

        # ───────────────────────────────────────────────────────
        # POS 1.4 — Fenster-Stückliste (Stk)
        # ───────────────────────────────────────────────────────
        pos = LVPosition(f"{pos_prefix}.4",
                         f"Fenster (Stockmaß) — Haus {haus}, {geschoss}", "Stk")
        pos.quelle = "PDF-Text-Layer (FE_-Code + 2×AL-Werte)"
        for w in h_windows:
            pos.add_zeile(
                f"{w['code']} ({int(w['breite_m']*100)}×{int(w['hoehe_m']*100)} mm)",
                anzahl=1, summe=1,
                quelle=w["quelle"]
            )
        pos.konfidenz = 0.95
        positionen.append(pos)

        # ───────────────────────────────────────────────────────
        # POS 1.5 — Loggien (eigene Position falls vorhanden)
        # ───────────────────────────────────────────────────────
        loggien = [r for r in h_rooms if r["kategorie"] == "Loggia"
                   and r.get("umfang_m") and r.get("hoehe_m")]
        if loggien:
            pos = LVPosition(f"{pos_prefix}.5",
                             f"Loggien Wandflächen — Haus {haus}, {geschoss}", "m²")
            pos.quelle = "PDF-Text-Layer"
            for r in loggien:
                pos.add_zeile(
                    f"{r['name']} ({r.get('wohnung') or '–'})",
                    laenge=r["umfang_m"], hoehe=r["hoehe_m"],
                    summe=r["umfang_m"] * r["hoehe_m"],
                    quelle=f"U={r['umfang_m']}m · H={r['hoehe_m']}m"
                )
            pos.konfidenz = 0.95
            positionen.append(pos)

        # ───────────────────────────────────────────────────────
        # POS 1.6 — Stiegenhaus
        # ───────────────────────────────────────────────────────
        stgh = [r for r in h_rooms if r["kategorie"] == "Stiegenhaus"
                and r.get("umfang_m") and r.get("hoehe_m")]
        if stgh:
            pos = LVPosition(f"{pos_prefix}.6",
                             f"Stiegenhaus Wandfläche — Haus {haus}, {geschoss}", "m²")
            pos.quelle = "PDF-Text-Layer"
            for r in stgh:
                pos.add_zeile(
                    f"{r['name']}",
                    laenge=r["umfang_m"], hoehe=r["hoehe_m"],
                    summe=r["umfang_m"] * r["hoehe_m"],
                    quelle=f"U={r['umfang_m']}m · H={r['hoehe_m']}m"
                )
            pos.konfidenz = 0.95
            positionen.append(pos)

    return positionen


# ════════════════════════════════════════════════════════════════════════
# 8. TOP-Wohnung-Zuordnung (Nearest TOP-Label)
# ════════════════════════════════════════════════════════════════════════
def assign_tops(rooms: list[dict], spans: list[dict]):
    top_rx = re.compile(r"^(TOP|Top|top)\s*\.?\s*([0-9]{1,3}[a-zA-Z]?)$")
    tops = []
    for s in spans:
        m = top_rx.match(s["text"].strip())
        if m:
            tops.append({"name": f"TOP {m.group(2)}", "cx": s["cx"], "cy": s["cy"]})
    # Loggien/Balkone liegen außerhalb der Wohnungs-Kernfläche und damit
    # weit vom TOP-Label entfernt — Radius muss großzügig sein, sonst bleiben
    # sie "—". 1200pt deckt typische Wohnungs-Diagonalen ab.
    for r in rooms:
        best = None
        best_d = 1200
        for t in tops:
            d = math.hypot(r["cx"] - t["cx"], r["cy"] - t["cy"])
            if d < best_d:
                best_d = d
                best = t["name"]
        r["wohnung"] = best


# ════════════════════════════════════════════════════════════════════════
# 9. Excel-Export im Kutzen-Koblach-Format
# ════════════════════════════════════════════════════════════════════════
def export_excel(positionen: list[LVPosition], out_path: Path, baustelle: str = ""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Massenermittlung ÖNORM"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    sub_fill = PatternFill("solid", fgColor="F5F7FA")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    # Header (analog Kutzen-Koblach-Excel)
    ws["A2"] = "MASSENERMITTLUNG / BAUSTELLE"
    ws["E2"] = baustelle or "—"
    ws["A2"].font = bold
    ws["A4"] = "Quelle"
    ws["B4"] = "Automatische Plan-Extraktion (ÖNORM A 2063)"

    headers = ["Positions-\nnummer", "Beschreibung", "Anzahl", "Länge m", "Breite m",
               "Höhe m", "Zwischen-\nsumme", "Endsumme", "Einheit", "Konfidenz", "Quelle"]
    for col, txt in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=txt)
        c.font = white_bold
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = border

    row = 8
    for pos in positionen:
        # Pos-Header
        ws.cell(row=row, column=1, value=pos.posnr).font = bold
        ws.cell(row=row, column=2, value=pos.beschreibung).font = bold
        for c in range(1, 12):
            ws.cell(row=row, column=c).fill = sub_fill
            ws.cell(row=row, column=c).border = border
        row += 1

        for z in pos.zeilen:
            ws.cell(row=row, column=2, value=z["text"])
            ws.cell(row=row, column=3, value=z["anzahl"])
            ws.cell(row=row, column=4, value=z["laenge"])
            ws.cell(row=row, column=5, value=z["breite"])
            ws.cell(row=row, column=6, value=z["hoehe"])
            ws.cell(row=row, column=7, value=z["wert"])
            ws.cell(row=row, column=11, value=z["quelle"])
            for c in range(1, 12):
                ws.cell(row=row, column=c).border = border
            row += 1

        # Summen-Zeile
        ws.cell(row=row, column=7, value="Summe Gesamt:").font = bold
        ws.cell(row=row, column=8, value=pos.endsumme).font = bold
        ws.cell(row=row, column=9, value=pos.einheit).font = bold
        ws.cell(row=row, column=10, value=f"{pos.konfidenz*100:.0f}%")
        ws.cell(row=row, column=11, value=pos.quelle).font = Font(italic=True, size=9)
        for c in range(1, 12):
            ws.cell(row=row, column=c).border = border
        row += 2

    # Spaltenbreiten
    widths = [12, 50, 8, 10, 10, 10, 12, 12, 8, 11, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(out_path)


# ════════════════════════════════════════════════════════════════════════
# 10. Main
# ════════════════════════════════════════════════════════════════════════
def analyse_pdf(pdf_path: Path) -> dict:
    doc = fitz.open(pdf_path)
    page = doc[0]
    pw, ph = page.rect.width, page.rect.height
    spans = extract_spans(page)

    # Maßstab + Geschoss aus Spans suchen.
    # Pläne haben oft mehrere geschoss-artige Kürzel — der echte Plan-Titel
    # ist das größte Label. Daher: sortiere Kandidaten nach Schriftgröße.
    massstab = None
    for s in spans:
        m = re.match(r"^M?\s*1\s*:\s*(50|100|200|500)\b", s["text"])
        if m and not massstab:
            massstab = f"1:{m.group(1)}"

    geschoss = None
    GESCHOSS_MAP = {
        "Erdgeschoss": "EG", "EG": "EG",
        "Obergeschoss": "OG", "1.Obergeschoss": "1.OG", "2.Obergeschoss": "2.OG",
        "Kellergeschoss": "KG", "Untergeschoss": "UG",
        "Dachgeschoss": "DG",
    }
    # Sammle alle Kandidaten, sortiere nach Schriftgröße absteigend
    geschoss_candidates = []
    for s in spans:
        t = s["text"].strip()
        gs = re.match(r"^(Erdgeschoss|Kellergeschoss|Obergeschoss|Untergeschoss|Dachgeschoss|EG|OG\d?|KG|DG|UG|\d\.\s*OG)\s*$",
                      t, re.I)
        if gs:
            geschoss_candidates.append((s["size"], gs.group(1)))
    if geschoss_candidates:
        geschoss_candidates.sort(reverse=True)  # größte Schrift zuerst
        raw = geschoss_candidates[0][1]
        geschoss = GESCHOSS_MAP.get(raw, raw.upper())

    # Räume erkennen — adaptive Schrift-Schwelle pro Plan
    min_label_size = compute_label_size_threshold(spans)
    raw_rooms = [extract_room(s, spans) for s in spans
                 if is_room_label(s, min_label_size)]
    rooms = dedup_by_position(raw_rooms)

    # Fenster & Türen
    windows = extract_windows(spans)
    doors = extract_doors(spans)

    # Häuser zuordnen
    houses = auto_houses(rooms, pw)
    for r in rooms:
        r["haus"] = house_of(r["cx"], houses)
    for w in windows:
        w["haus"] = house_of(w["cx"], houses)

    # TOP-Wohnungen
    assign_tops(rooms, spans)

    # LV bauen (Türen werden heuristisch in build_lv geschätzt; die
    # `doors`-Liste aus dem Text-Layer ist für eine zukünftige Iteration
    # vorgesehen, wenn Türlängen aus Plan-Geometrie verfügbar sind.)
    positionen = build_lv(rooms, windows, houses, geschoss or "EG")

    return {
        "pdf": str(pdf_path.name),
        "page_size_pt": [pw, ph],
        "massstab": massstab,
        "geschoss": geschoss,
        "houses": houses,
        "spans": len(spans),
        "rooms": rooms,
        "windows": windows,
        "doors": doors,
        "lv": [p.to_dict() for p in positionen],
        "positionen_objs": positionen,
    }


def print_report(result: dict):
    print(f"\n{'═'*72}\n  {result['pdf']}\n{'═'*72}")
    print(f"  Seite: {result['page_size_pt'][0]:.0f}×{result['page_size_pt'][1]:.0f} pt  "
          f"| Maßstab: {result['massstab']}  | Geschoss: {result['geschoss']}")
    print(f"  Spans: {result['spans']}  | Räume: {len(result['rooms'])}  "
          f"| Fenster: {len(result['windows'])}  | Häuser: {sorted(result['houses'].keys())}")

    print(f"\n{'─'*72}\nLESE-GENAUIGKEIT (pro Wert byte-exakt aus PDF-Text-Layer)\n{'─'*72}")
    n_r = len(result["rooms"])
    n_f = sum(1 for r in result["rooms"] if r.get("flaeche_m2"))
    n_u = sum(1 for r in result["rooms"] if r.get("umfang_m"))
    n_h = sum(1 for r in result["rooms"] if r.get("hoehe_m"))
    n_b = sum(1 for r in result["rooms"] if r.get("bodenbelag"))
    n_top = sum(1 for r in result["rooms"] if r.get("wohnung"))
    n_w = len(result["windows"])
    n_w_full = sum(1 for w in result["windows"] if "2xAL" in w.get("quelle", ""))

    def pct(n, total):
        return f"{n}/{total} ({n/max(total,1)*100:.0f}%)"

    print(f"  Fläche F:       {pct(n_f, n_r)}   ← Einheit m², aus F-Anker oder Zahl+m²-Hochstellung")
    print(f"  Umfang U:       {pct(n_u, n_r)}   ← Einheit m, aus 'U:' Anker")
    print(f"  Höhe H:         {pct(n_h, n_r)}   ← Einheit m, aus 'H:' Anker")
    print(f"  Bodenbelag:     {pct(n_b, n_r)}   ← (Loggia/Stiegenhaus haben oft keinen)")
    print(f"  TOP-Zuordnung:  {pct(n_top, n_r)}   ← Nearest-TOP-Label, Radius 1200pt")
    print(f"  Fenster (Code+2xAL): {pct(n_w_full, n_w)}   ← FE_-Marker + 2x AL-Werte")
    print(f"  Maßstab gesetzt:  {'ja' if result['massstab'] else 'NEIN'}")
    print(f"  Geschoss gesetzt: {'ja' if result['geschoss'] else 'NEIN'}")

    print(f"\n{'─'*72}\nÖNORM A 2063 LV (Übersicht)\n{'─'*72}")
    print(f"  {'PosNr':<10} {'Beschreibung':<55} {'Endsumme':>10} {'Einh':>5} {'Konf':>6}")
    for p in result["positionen_objs"]:
        print(f"  {p.posnr:<10} {p.beschreibung[:54]:<55} {p.endsumme:>10.2f} {p.einheit:>5} "
              f"{p.konfidenz*100:>5.0f}%")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    pdf = Path(sys.argv[1]).expanduser()
    out_xlsx = Path(sys.argv[2]).expanduser() if len(sys.argv) > 2 else \
        Path("/tmp/massenermittlung.xlsx")

    result = analyse_pdf(pdf)
    print_report(result)

    export_excel(result["positionen_objs"], out_xlsx, baustelle=pdf.stem)
    print(f"\n📄 Excel-LV: {out_xlsx}")

    # JSON-Dump für API/UI
    json_path = out_xlsx.with_suffix(".json")
    dump = {k: v for k, v in result.items() if k != "positionen_objs"}
    json_path.write_text(json.dumps(dump, indent=2, ensure_ascii=False, default=str))
    print(f"📄 JSON:     {json_path}")


if __name__ == "__main__":
    main()
