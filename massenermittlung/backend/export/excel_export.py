"""
ÖNORM-style Excel export for Massenermittlung data.

Generates a professional multi-sheet Excel workbook with formatted headers,
confidence-based cell coloring, and summary rows.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CONFIDENCE_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CONFIDENCE_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CONFIDENCE_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SUMMARY_FONT = Font(bold=True, size=11)
SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _apply_header_style(ws, columns: list[str]) -> None:
    """Write header row and apply formatting."""
    for col_idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_column_width(ws, *, min_width: int = 10, max_width: int = 50) -> None:
    """Adjust column widths based on cell content length."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                # For multi-line cells take the longest line
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                max_len = max(max_len, longest + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def _confidence_fill(value: Any) -> PatternFill | None:
    """Return the appropriate fill for a confidence value."""
    try:
        conf = float(value)
    except (TypeError, ValueError):
        return None
    if conf >= 80:
        return CONFIDENCE_GREEN
    if conf >= 60:
        return CONFIDENCE_YELLOW
    return CONFIDENCE_RED


def _safe_get(obj: dict, key: str, default: Any = "") -> Any:
    """Safely retrieve a value from a dict."""
    if isinstance(obj, dict):
        return obj.get(key, default)
    return default


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_massenermittlung_sheet(ws, massen: list[dict]) -> None:
    """Populate the Massenermittlung sheet."""
    columns = [
        "Pos-Nr",
        "Beschreibung",
        "Raum",
        "Berechnung",
        "Endsumme",
        "Einheit",
        "Gewerk",
        "Konfidenz",
    ]
    _apply_header_style(ws, columns)

    total_endsumme = 0.0

    for row_idx, masse in enumerate(massen, start=2):
        ws.cell(row=row_idx, column=1, value=_safe_get(masse, "pos_nr")).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=_safe_get(masse, "beschreibung")).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=_safe_get(masse, "raum")).border = THIN_BORDER

        berechnung_cell = ws.cell(row=row_idx, column=4, value=_safe_get(masse, "berechnung"))
        berechnung_cell.alignment = Alignment(wrap_text=True, vertical="top")
        berechnung_cell.border = THIN_BORDER

        endsumme = _safe_get(masse, "endsumme", 0)
        try:
            endsumme = float(endsumme)
            total_endsumme += endsumme
        except (TypeError, ValueError):
            endsumme = 0
        ws.cell(row=row_idx, column=5, value=endsumme).border = THIN_BORDER
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00'

        ws.cell(row=row_idx, column=6, value=_safe_get(masse, "einheit")).border = THIN_BORDER
        ws.cell(row=row_idx, column=7, value=_safe_get(masse, "gewerk")).border = THIN_BORDER

        konfidenz = _safe_get(masse, "konfidenz", 0)
        conf_cell = ws.cell(row=row_idx, column=8, value=konfidenz)
        conf_cell.border = THIN_BORDER
        conf_fill = _confidence_fill(konfidenz)
        if conf_fill:
            conf_cell.fill = conf_fill

    # Summary row
    summary_row = len(massen) + 2
    for col in range(1, 9):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = SUMMARY_FONT
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    ws.cell(row=summary_row, column=1, value="GESAMT")
    ws.cell(row=summary_row, column=5, value=total_endsumme).number_format = '#,##0.00'

    # Freeze header
    ws.freeze_panes = "A2"


def _build_raeume_sheet(ws, raeume: list[dict]) -> None:
    """Populate the Raeume sheet."""
    columns = [
        "Name",
        "Bodenbelag",
        "Fläche m\u00b2",
        "Umfang m",
        "Höhe m",
        "Wandfläche m\u00b2",
    ]
    _apply_header_style(ws, columns)

    total_flaeche = 0.0
    total_wandflaeche = 0.0

    for row_idx, raum in enumerate(raeume, start=2):
        ws.cell(row=row_idx, column=1, value=_safe_get(raum, "name")).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=_safe_get(raum, "bodenbelag")).border = THIN_BORDER

        flaeche = _safe_get(raum, "flaeche", 0)
        try:
            flaeche = float(flaeche)
            total_flaeche += flaeche
        except (TypeError, ValueError):
            flaeche = 0
        ws.cell(row=row_idx, column=3, value=flaeche).border = THIN_BORDER
        ws.cell(row=row_idx, column=3).number_format = '#,##0.00'

        umfang = _safe_get(raum, "umfang", 0)
        try:
            umfang = float(umfang)
        except (TypeError, ValueError):
            umfang = 0
        ws.cell(row=row_idx, column=4, value=umfang).border = THIN_BORDER
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00'

        hoehe = _safe_get(raum, "hoehe", 0)
        try:
            hoehe = float(hoehe)
        except (TypeError, ValueError):
            hoehe = 0
        ws.cell(row=row_idx, column=5, value=hoehe).border = THIN_BORDER
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00'

        wandflaeche = _safe_get(raum, "wandflaeche", 0)
        try:
            wandflaeche = float(wandflaeche)
            total_wandflaeche += wandflaeche
        except (TypeError, ValueError):
            wandflaeche = 0
        ws.cell(row=row_idx, column=6, value=wandflaeche).border = THIN_BORDER
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'

    # Summary row
    summary_row = len(raeume) + 2
    for col in range(1, 7):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = SUMMARY_FONT
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    ws.cell(row=summary_row, column=1, value="GESAMT")
    ws.cell(row=summary_row, column=3, value=total_flaeche).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=6, value=total_wandflaeche).number_format = '#,##0.00'

    ws.freeze_panes = "A2"


def _build_fenster_sheet(ws, fenster: list[dict]) -> None:
    """Populate the Fenster sheet."""
    columns = [
        "Bezeichnung",
        "Raum",
        "AL Breite mm",
        "AL Höhe mm",
        "RB Breite mm",
        "RB Höhe mm",
        "Fläche m\u00b2",
    ]
    _apply_header_style(ws, columns)

    total_flaeche = 0.0

    for row_idx, f in enumerate(fenster, start=2):
        ws.cell(row=row_idx, column=1, value=_safe_get(f, "bezeichnung")).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=_safe_get(f, "raum")).border = THIN_BORDER

        for col, key in [(3, "al_breite"), (4, "al_hoehe"), (5, "rb_breite"), (6, "rb_hoehe")]:
            val = _safe_get(f, key, 0)
            try:
                val = float(val)
            except (TypeError, ValueError):
                val = 0
            ws.cell(row=row_idx, column=col, value=val).border = THIN_BORDER
            ws.cell(row=row_idx, column=col).number_format = '#,##0'

        flaeche = _safe_get(f, "flaeche", 0)
        try:
            flaeche = float(flaeche)
            total_flaeche += flaeche
        except (TypeError, ValueError):
            flaeche = 0
        ws.cell(row=row_idx, column=7, value=flaeche).border = THIN_BORDER
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'

    # Summary row
    summary_row = len(fenster) + 2
    for col in range(1, 8):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = SUMMARY_FONT
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    ws.cell(row=summary_row, column=1, value="GESAMT")
    ws.cell(row=summary_row, column=7, value=total_flaeche).number_format = '#,##0.00'

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel(
    massen: list[dict],
    raeume: list[dict],
    fenster: list[dict],
    projekt_name: str,
) -> bytes:
    """
    Generate an ÖNORM-style Excel workbook for a Massenermittlung project.

    Args:
        massen: List of dicts with keys: pos_nr, beschreibung, raum,
                berechnung, endsumme, einheit, gewerk, konfidenz.
        raeume: List of dicts with keys: name, bodenbelag, flaeche,
                umfang, hoehe, wandflaeche.
        fenster: List of dicts with keys: bezeichnung, raum, al_breite,
                 al_hoehe, rb_breite, rb_hoehe, flaeche.
        projekt_name: Name of the project (used in the workbook title).

    Returns:
        Excel file content as bytes, ready for streaming or saving.
    """
    wb = Workbook()

    # Sheet 1 – Massenermittlung
    ws_massen = wb.active
    ws_massen.title = "Massenermittlung"
    _build_massenermittlung_sheet(ws_massen, massen)
    _auto_column_width(ws_massen)

    # Sheet 2 – Räume
    ws_raeume = wb.create_sheet("Räume")
    _build_raeume_sheet(ws_raeume, raeume)
    _auto_column_width(ws_raeume)

    # Sheet 3 – Fenster
    ws_fenster = wb.create_sheet("Fenster")
    _build_fenster_sheet(ws_fenster, fenster)
    _auto_column_width(ws_fenster)

    # Set workbook properties
    wb.properties.title = f"Massenermittlung – {projekt_name}"
    wb.properties.creator = "Massenermittlung App"

    # Write to BytesIO and return bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
